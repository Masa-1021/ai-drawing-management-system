"""Excelパースサービス

摘番マスタExcel、摘要表Excelのパース処理
"""
from typing import List, Optional, Tuple, Dict
from pydantic import BaseModel
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import logging

logger = logging.getLogger(__name__)


class SpecNumberRow(BaseModel):
    """摘番マスタ1行"""
    spec_number: str
    title: Optional[str] = None
    model_name: Optional[str] = None
    material_code: Optional[str] = None
    usage_location: Optional[str] = None
    line_name: Optional[str] = None
    equipment_name: Optional[str] = None
    design_date: Optional[str] = None
    designer: Optional[str] = None
    reference_drawing: Optional[str] = None
    remarks: Optional[str] = None


class SpecSheetHeader(BaseModel):
    """摘要表ヘッダー情報"""
    equipment_name: Optional[str] = None
    line_name: Optional[str] = None
    model_name: Optional[str] = None
    spec_number: str
    order_number: Optional[str] = None
    created_by: Optional[str] = None
    checked_by: Optional[str] = None
    designed_by: Optional[str] = None
    approved_by: Optional[str] = None


class SpecSheetRevisionRow(BaseModel):
    """摘要表改定履歴"""
    revision_symbol: str
    revision_date: Optional[str] = None
    description: Optional[str] = None
    created_by: Optional[str] = None
    checked_by: Optional[str] = None
    approved_by: Optional[str] = None
    remarks: Optional[str] = None


class SpecSheetItemRow(BaseModel):
    """摘要表部品1行"""
    row_number: int
    part_name: Optional[str] = None
    drawing_number: Optional[str] = None
    sub_number: Optional[str] = None
    item_number: Optional[str] = None
    material: Optional[str] = None
    heat_treatment: Optional[str] = None
    surface_treatment: Optional[str] = None
    quantity_per_set: Optional[int] = None
    required_quantity: Optional[int] = None
    revision: Optional[str] = None
    part_type: str  # 'assembly', 'unit', 'part', 'purchased'
    parent_name: Optional[str] = None  # 列27の値


class ExcelParserService:
    """Excelパースサービス"""

    def __init__(self, config: Dict):
        self.config = config
        self.spec_sheet_config = config.get("specSheetParsing", {})
        self.spec_number_config = config.get("specNumberParsing", {})

    def parse_spec_number_excel(self, file_path: str) -> List[SpecNumberRow]:
        """摘番一括検索Excelをパース

        Args:
            file_path: Excelファイルパス

        Returns:
            摘番マスタ行のリスト
        """
        logger.info(f"摘番マスタExcelパース開始: {file_path}")

        wb = load_workbook(file_path, data_only=True)
        sheet_name = self.spec_number_config.get("sheetName", "検索用")

        if sheet_name not in wb.sheetnames:
            # シート名が見つからない場合、最初のシートを使用
            logger.warning(f"シート '{sheet_name}' が見つかりません。最初のシートを使用します。")
            ws = wb.active
        else:
            ws = wb[sheet_name]

        col_mapping = self.spec_number_config.get("columnMapping", {
            "spec_number": 2,
            "title": 3,
            "model_name": 4,
            "material_code": 5,
            "usage_location": 6,
            "line_name": 7,
            "equipment_name": 8,
            "design_date": 9,
            "designer": 10,
            "reference_drawing": 11,
            "remarks": 12
        })
        start_row = self.spec_number_config.get("dataStartRow", 4)

        results = []
        for row in range(start_row, ws.max_row + 1):
            spec_number = self._get_cell(ws, row, col_mapping.get("spec_number", 2))
            if not spec_number:
                continue

            try:
                results.append(SpecNumberRow(
                    spec_number=spec_number,
                    title=self._get_cell(ws, row, col_mapping.get("title", 3)),
                    model_name=self._get_cell(ws, row, col_mapping.get("model_name", 4)),
                    material_code=self._get_cell(ws, row, col_mapping.get("material_code", 5)),
                    usage_location=self._get_cell(ws, row, col_mapping.get("usage_location", 6)),
                    line_name=self._get_cell(ws, row, col_mapping.get("line_name", 7)),
                    equipment_name=self._get_cell(ws, row, col_mapping.get("equipment_name", 8)),
                    design_date=self._get_cell(ws, row, col_mapping.get("design_date", 9)),
                    designer=self._get_cell(ws, row, col_mapping.get("designer", 10)),
                    reference_drawing=self._get_cell(ws, row, col_mapping.get("reference_drawing", 11)),
                    remarks=self._get_cell(ws, row, col_mapping.get("remarks", 12)),
                ))
            except Exception as e:
                logger.warning(f"行 {row} のパースエラー: {e}")
                continue

        wb.close()
        logger.info(f"摘番マスタExcelパース完了: {len(results)}件")
        return results

    def parse_spec_sheet_excel(
        self, file_path: str
    ) -> Tuple[SpecSheetHeader, List[SpecSheetRevisionRow], List[SpecSheetItemRow]]:
        """摘要表Excelをパース

        Args:
            file_path: Excelファイルパス

        Returns:
            (ヘッダー情報, 改定履歴リスト, 部品リスト)
        """
        logger.info(f"摘要表Excelパース開始: {file_path}")

        wb = load_workbook(file_path, data_only=True)

        # 表紙シートからヘッダー抽出
        cover_sheet_name = self.spec_sheet_config.get("coverSheetName", "表紙")
        if cover_sheet_name in wb.sheetnames:
            cover_ws = wb[cover_sheet_name]
            header = self._parse_cover_sheet(cover_ws)
            revisions = self._parse_revisions(cover_ws)
        else:
            # 表紙シートがない場合、最初のシートから抽出を試みる
            logger.warning(f"シート '{cover_sheet_name}' が見つかりません。")
            header = SpecSheetHeader(spec_number="UNKNOWN")
            revisions = []

        # 摘要表シートから部品リスト抽出
        items_sheet_name = self.spec_sheet_config.get("itemsSheetName", "摘要表")
        if items_sheet_name in wb.sheetnames:
            items_ws = wb[items_sheet_name]
            items = self._parse_items_sheet(items_ws)
        else:
            logger.warning(f"シート '{items_sheet_name}' が見つかりません。")
            items = []

        wb.close()
        logger.info(f"摘要表Excelパース完了: 改定{len(revisions)}件, 部品{len(items)}件")
        return header, revisions, items

    def _parse_cover_sheet(self, ws: Worksheet) -> SpecSheetHeader:
        """表紙シートからヘッダー情報抽出"""
        header_mapping = self.spec_sheet_config.get("headerMapping", {})

        def get_mapped_cell(field_name: str, default_row: int, default_col: int) -> Optional[str]:
            mapping = header_mapping.get(field_name, {"row": default_row, "col": default_col})
            return self._get_cell(ws, mapping.get("row", default_row), mapping.get("col", default_col))

        spec_number = get_mapped_cell("spec_number", 4, 11) or "UNKNOWN"

        return SpecSheetHeader(
            equipment_name=get_mapped_cell("equipment_name", 2, 3),
            line_name=get_mapped_cell("line_name", 3, 3),
            model_name=get_mapped_cell("model_name", 5, 3),
            spec_number=spec_number,
            order_number=get_mapped_cell("order_number", 4, 12),
            created_by=get_mapped_cell("created_by", 4, 6),
            checked_by=get_mapped_cell("checked_by", 4, 7),
            designed_by=get_mapped_cell("designed_by", 4, 8),
            approved_by=get_mapped_cell("approved_by", 4, 9),
        )

    def _parse_revisions(self, ws: Worksheet) -> List[SpecSheetRevisionRow]:
        """表紙シートから改定履歴抽出

        改定履歴は通常、表紙の下部にあるテーブルから抽出
        """
        revisions = []

        # 改定履歴テーブルを探す（一般的な位置: 行10以降）
        # 実際のファイル構造に応じて調整が必要
        for row in range(10, min(ws.max_row + 1, 30)):
            # 改定記号を探す（A, B, C...の単一文字）
            symbol = self._get_cell(ws, row, 1)
            if symbol and len(symbol) == 1 and symbol.isalpha():
                revisions.append(SpecSheetRevisionRow(
                    revision_symbol=symbol,
                    revision_date=self._get_cell(ws, row, 2),
                    description=self._get_cell(ws, row, 3),
                    created_by=self._get_cell(ws, row, 4),
                    checked_by=self._get_cell(ws, row, 5),
                    approved_by=self._get_cell(ws, row, 6),
                    remarks=self._get_cell(ws, row, 7),
                ))

        return revisions

    def _parse_items_sheet(self, ws: Worksheet) -> List[SpecSheetItemRow]:
        """摘要表シートから部品リスト抽出"""
        col_mapping = self.spec_sheet_config.get("columnMapping", {
            "revision": 5,
            "row_number": 6,
            "part_name": 7,
            "drawing_number": 8,
            "sub_number": 9,
            "item_number": 10,
            "material": 11,
            "heat_treatment": 12,
            "surface_treatment": 13,
            "quantity_per_set": 14,
            "required_quantity": 15,
            "parent_name": 27
        })
        start_row = self.spec_sheet_config.get("itemsStartRow", 7)

        results = []
        for row in range(start_row, ws.max_row + 1):
            row_number_val = ws.cell(row, col_mapping.get("row_number", 6)).value
            if row_number_val is None:
                continue

            try:
                row_number = int(row_number_val)
            except (ValueError, TypeError):
                continue

            part_name = self._get_cell(ws, row, col_mapping.get("part_name", 7))
            drawing_number = self._get_cell(ws, row, col_mapping.get("drawing_number", 8))
            parent_name = self._get_cell(ws, row, col_mapping.get("parent_name", 27))

            # 部品タイプ判定
            part_type = self._determine_part_type(part_name, drawing_number, parent_name)

            results.append(SpecSheetItemRow(
                row_number=row_number,
                part_name=part_name,
                drawing_number=drawing_number,
                sub_number=self._get_cell(ws, row, col_mapping.get("sub_number", 9)),
                item_number=self._get_cell(ws, row, col_mapping.get("item_number", 10)),
                material=self._get_cell(ws, row, col_mapping.get("material", 11)),
                heat_treatment=self._get_cell(ws, row, col_mapping.get("heat_treatment", 12)),
                surface_treatment=self._get_cell(ws, row, col_mapping.get("surface_treatment", 13)),
                quantity_per_set=self._get_int(ws, row, col_mapping.get("quantity_per_set", 14)),
                required_quantity=self._get_int(ws, row, col_mapping.get("required_quantity", 15)),
                revision=self._get_cell(ws, row, col_mapping.get("revision", 5)),
                part_type=part_type,
                parent_name=parent_name,
            ))

        return results

    def _determine_part_type(
        self,
        part_name: Optional[str],
        drawing_number: Optional[str],
        parent_name: Optional[str]
    ) -> str:
        """部品タイプを判定

        判定ロジック:
        1. 品名にASSY, 組立, 総組などが含まれる → assembly
        2. 品名にUNIT, ユニットなどが含まれる → unit
        3. 図面番号がない → purchased
        4. 列27が「組図」→ assembly（トップレベル）
        5. その他 → part

        Args:
            part_name: 品名
            drawing_number: 図面番号
            parent_name: 親ユニット名（列27の値）

        Returns:
            部品タイプ ('assembly', 'unit', 'part', 'purchased')
        """
        keywords = self.spec_sheet_config.get("partTypeKeywords", {
            "assembly": ["ASSY", "組立", "総組", "ASSEMBLY"],
            "unit": ["UNIT", "ユニット"]
        })

        if not part_name:
            return "purchased" if not drawing_number else "part"

        name_upper = part_name.upper()

        # 品名から判定
        for kw in keywords.get("assembly", []):
            if kw.upper() in name_upper:
                return "assembly"

        for kw in keywords.get("unit", []):
            if kw.upper() in name_upper:
                return "unit"

        # 図面番号がなければ購入品
        if not drawing_number:
            return "purchased"

        # 列27が「組図」ならトップレベル組図
        if parent_name and parent_name == "組図":
            return "assembly"

        return "part"

    def _get_cell(self, ws: Worksheet, row: int, col: int) -> Optional[str]:
        """セル値を文字列で取得"""
        val = ws.cell(row, col).value
        if val is None:
            return None
        return str(val).strip()

    def _get_int(self, ws: Worksheet, row: int, col: int) -> Optional[int]:
        """セル値を整数で取得"""
        val = ws.cell(row, col).value
        if val is None:
            return None
        try:
            return int(val)
        except (ValueError, TypeError):
            return None
